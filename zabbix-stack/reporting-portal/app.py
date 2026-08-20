#!/usr/bin/env python3
from flask import Flask, render_template_string, request, send_file, jsonify
import requests, io, os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from prometheus_flask_exporter import PrometheusMetrics

app = Flask(__name__)
metrics = PrometheusMetrics(app)
# static information as metric
metrics.info('app_info', 'Application info', version='1.0.0')

ZABBIX_URL = os.environ.get('ZABBIX_URL', 'http://zabbix-web:8080/api_jsonrpc.php')
ZABBIX_USER = os.environ.get('ZABBIX_USER', 'Admin')
ZABBIX_PASSWORD = os.environ.get('ZABBIX_PASSWORD', 'zabbix')
ZABBIX_API_TOKEN = os.environ.get('ZABBIX_API_TOKEN', '')

class ZabbixAPI:
    def __init__(self):
        self.auth = None

    def login(self):
        # Prefer the long-lived API token (same one the other zabbix-stack
        # scripts use) over the Admin password, which has proven unreliable
        # (rotates / triggers Zabbix's failed-login lockout).
        if ZABBIX_API_TOKEN:
            self.auth = ZABBIX_API_TOKEN
            return True
        try:
            r = requests.post(ZABBIX_URL, json={
                "jsonrpc": "2.0",
                "method": "user.login",
                "params": {"username": ZABBIX_USER, "password": ZABBIX_PASSWORD},
                "id": 1
            }, timeout=10)
            self.auth = r.json().get('result')
            return bool(self.auth)
        except:
            return False

    def call(self, method, params):
        if not self.auth:
            self.login()

        headers = {'Authorization': f'Bearer {self.auth}'} if self.auth else {}
        try:
            r = requests.post(ZABBIX_URL, json={
                "jsonrpc": "2.0",
                "method": method,
                "params": params,
                "id": 1
            }, headers=headers, timeout=30)

            res = r.json()
            if 'error' in res and (res['error']['code'] == -32602 or 'auth' in res['error']['message'].lower()):
                r = requests.post(ZABBIX_URL, json={
                    "jsonrpc": "2.0",
                    "method": method,
                    "params": params,
                    "auth": self.auth,
                    "id": 1
                }, timeout=30)
                res = r.json()

            return res.get('result', [])
        except Exception as e:
            print(f"API Call Error: {e}")
            return []

    def get_detailed_inventory(self):
        return self.call("host.get", {
            "output": ["hostid", "host", "name", "status"],
            "selectGroups": ["name"],
            "selectInterfaces": "extend",
            "selectInventory": "extend"
        })

    def get_active_problems(self):
        return self.call("trigger.get", {
            "output": ["triggerid", "description", "priority", "lastchange"],
            "selectHosts": ["name"],
            "only_true": True,
            "monitored": True,
            "sortfield": "priority",
            "sortorder": "DESC"
        })

    def get_triggers(self):
        return self.call("trigger.get", {
            "output": ["triggerid", "description", "priority", "value", "lastchange"],
            "selectHosts": ["name"],
            "active": True,
            "monitored": True
        })

    def get_native_slas(self):
        slas = self.call("sla.get", {"output": "extend"})
        if not slas: return []

        # Get all services mapping for name lookup
        all_services = {s['serviceid']: s['name'] for s in self.call("service.get", {"output": ["serviceid", "name"]})}

        results = []
        for sla in slas:
            period_results = self.call("sla.getsli", {
                "slaid": sla['slaid'],
                "period_from": int((datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)).timestamp())
            })

            if period_results and 'sli' in period_results and 'serviceids' in period_results:
                # Zabbix 7.x returns sli as a list of lists: sli[period_index][service_index]
                # We typically want the most recent period (the last one or first if only one)
                for period_idx, period_sli in enumerate(period_results['sli']):
                    for svc_idx, sli_data in enumerate(period_sli):
                        svc_id = str(period_results['serviceids'][svc_idx])
                        svc_name = all_services.get(svc_id, f"Service {svc_id}")

                        actual_sli = sli_data.get('sli')
                        results.append({
                            "SLA Name": sla['name'],
                            "Service": svc_name,
                            "Target": f"{sla['slo']}%",
                            "Actual": f"{round(float(actual_sli), 3)}%" if actual_sli is not None else "N/A",
                            "Status": "OK" if (actual_sli or 0) >= float(sla['slo']) else "FAIL"
                        })
        return results

    def get_top_triggers(self):
        # Fetch events for the last 7 days to count frequency
        seven_days_ago = int((datetime.now() - timedelta(days=7)).timestamp())
        events = self.call("event.get", {
            "output": ["objectid"],
            "time_from": seven_days_ago,
            "source": 0, # EventSource.TRIGGERS
            "object": 0, # EventObject.TRIGGER
            "value": 1,   # Problem
        })

        if not events: return []

        # Count frequencies
        counts = {}
        for e in events:
            tid = e['objectid']
            counts[tid] = counts.get(tid, 0) + 1

        # Sort and take top 100
        top_tids = sorted(counts.items(), key=lambda x: x[1], reverse=True)[:100]
        if not top_tids: return []

        # Fetch details for these triggers
        trigger_details = self.call("trigger.get", {
            "output": ["triggerid", "description", "priority"],
            "triggerids": [x[0] for x in top_tids],
            "selectHosts": ["name"]
        })

        detail_map = {t['triggerid']: t for t in trigger_details}

        results = []
        sev_map = {0: 'Not classified', 1: 'Info', 2: 'Warning', 3: 'Average', 4: 'High', 5: 'Disaster'}

        for tid, count in top_tids:
            t = detail_map.get(tid)
            if t:
                results.append({
                    "Trigger ID": tid,
                    "Host": t['hosts'][0]['name'] if t.get('hosts') else 'N/A',
                    "Description": t['description'],
                    "Severity": sev_map.get(int(t['priority']), 'Unknown'),
                    "Frequency (Last 7 Days)": count
                })
        return results

    def get_availability_data(self):
        hosts = self.call("host.get", {
            "output": ["hostid", "host", "name", "status"],
            "selectInterfaces": "extend"
        })
        data = []
        status_map = {'0': 'Unknown', '1': 'Available', '2': 'Unavailable'}

        for h in hosts:
            agent_status = "N/A"
            snmp_status = "N/A"
            errors = []
            interfaces = h.get('interfaces', [])
            for iface in interfaces:
                itype = str(iface.get('type'))
                iavail = str(iface.get('available', '0'))
                ierror = iface.get('error', '')
                if itype == '1':
                    agent_status = status_map.get(iavail, 'Unknown')
                    if ierror: errors.append(f"Agent: {ierror}")
                elif itype == '2':
                    snmp_status = status_map.get(iavail, 'Unknown')
                    if ierror: errors.append(f"SNMP: {ierror}")

            uptime = "0%"
            if agent_status == 'Available' or snmp_status == 'Available': uptime = "100%"
            elif agent_status == 'N/A' and snmp_status == 'N/A': uptime = "Unknown"
            if h.get('status') == '1': uptime = "Disabled"

            data.append({
                'Host': h.get('name', h.get('host')),
                'IP': (interfaces[0].get('ip', 'N/A') if interfaces else 'N/A'),
                'Agent Status': agent_status,
                'SNMP Status': snmp_status,
                'Estimated Uptime': uptime,
                'Status Detail': "; ".join(errors) if errors else "OK"
            })
        return data

zabbix = ZabbixAPI()

def format_excel_sheet(ws, data, title):
    if not data:
        ws['A1'] = "No data available"
        return

    cols = list(data[0].keys())
    hdr_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")

    for i, c in enumerate(cols, 1):
        cell = ws.cell(1, i, c.replace('_', ' ').title())
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')

    for r, row in enumerate(data, 2):
        for c, col in enumerate(cols, 1):
            v = row.get(col, '')
            if col in ['clock', 'lastchange', 'Time', 'Last Change'] and v:
                try: v = datetime.fromtimestamp(int(v)).strftime('%Y-%m-%d %H:%M')
                except: pass
            cell = ws.cell(r, c, v)
            val_str = str(v)
            if val_str in ['Available', 'OK', '100%']: cell.font = Font(color="008000")
            elif val_str in ['Unavailable', '0%', 'PROBLEM', 'FAIL']: cell.font = Font(color="FF0000")

    for i in range(1, len(cols) + 1):
        width = 25
        if cols[i-1] in ['Status Detail', 'Os', 'Hardware']: width = 40
        ws.column_dimensions[get_column_letter(i)].width = width

def make_excel(data, title):
    wb = Workbook()
    format_excel_sheet(wb.active, data, title)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

HTML = """<!DOCTYPE html><html lang="en"><head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Infrastructure Reports</title>
<style>
    :root{
        --page:#f8fafc; --surface:#ffffff;
        --ink:#0f172a; --ink-2:#475569; --ink-3:#94a3b8;
        --border:#e2e8f0; --border-strong:#cbd5e1;
        --accent:#2563eb; --accent-hover:#1d4ed8; --accent-ink:#1e40af;
        --accent-wash:#eff6ff;
        --critical:#dc2626; --good:#16a34a;
        --shadow:0 1px 2px rgba(15,23,42,.05);
        --shadow-md:0 4px 12px rgba(15,23,42,.07);
    }
    @media (prefers-color-scheme: dark){
        :root{
            --page:#0f172a; --surface:#1e293b;
            --ink:#f1f5f9; --ink-2:#cbd5e1; --ink-3:#7c8aa0;
            --border:#2c3a50; --border-strong:#3b4c66;
            --accent:#3b82f6; --accent-hover:#60a5fa; --accent-ink:#93c5fd;
            --accent-wash:rgba(59,130,246,.12);
            --critical:#f87171; --good:#4ade80;
            --shadow:none; --shadow-md:none;
        }
    }
    *{margin:0;padding:0;box-sizing:border-box}
    body{font-family:Inter,-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;
         background:var(--page);color:var(--ink);font-size:15px;line-height:1.55;
         -webkit-font-smoothing:antialiased}
    .wrap{max-width:1080px;margin:0 auto;padding:0 24px 64px}

    header{background:var(--surface);border-bottom:1px solid var(--border)}
    .head-inner{max-width:1080px;margin:0 auto;padding:16px 24px;display:flex;align-items:center;justify-content:space-between;gap:16px;flex-wrap:wrap}
    .brand{display:flex;align-items:center;gap:11px}
    .brand-mark{width:34px;height:34px;border-radius:8px;background:var(--accent);display:flex;align-items:center;justify-content:center;flex:none}
    .brand-mark svg{width:19px;height:19px;color:#fff}
    .brand .name{font-weight:650;font-size:15.5px;letter-spacing:-.01em}
    .brand .sub{font-size:12px;color:var(--ink-3);margin-top:-2px}
    .head-meta{display:flex;align-items:center;gap:12px;font-size:13px;color:var(--ink-3)}
    .live{display:inline-flex;align-items:center;gap:7px;padding:4px 12px;background:var(--page);border:1px solid var(--border);border-radius:999px;color:var(--ink-2);font-size:12.5px;font-weight:500}
    .live .dot{width:7px;height:7px;border-radius:50%;background:var(--good)}

    .hero{padding:36px 0 4px}
    .hero h1{font-size:clamp(24px,3.5vw,30px);font-weight:700;letter-spacing:-.02em;line-height:1.2}
    .hero p{color:var(--ink-2);margin-top:8px;max-width:60ch;font-size:15px}

    .kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:14px;margin:26px 0 4px}
    .kpi{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:16px 18px;box-shadow:var(--shadow)}
    .kpi .label{font-size:12.5px;font-weight:500;color:var(--ink-2);display:flex;align-items:center;gap:7px}
    .kpi .label .sdot{width:8px;height:8px;border-radius:50%;flex:none}
    .kpi .value{font-size:30px;font-weight:650;letter-spacing:-.02em;margin-top:4px}
    .kpi .note{font-size:12px;color:var(--ink-3);margin-top:1px}

    h2.section{font-size:16px;font-weight:650;letter-spacing:-.01em;margin:38px 0 14px}

    .reports{display:grid;grid-template-columns:repeat(auto-fill,minmax(310px,1fr));gap:14px}
    .report{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:20px;box-shadow:var(--shadow);
            display:flex;flex-direction:column;gap:10px;transition:box-shadow .15s, border-color .15s}
    .report:hover{border-color:var(--border-strong);box-shadow:var(--shadow-md)}
    .report .top{display:flex;align-items:flex-start;justify-content:space-between;gap:10px}
    .report .icon{width:38px;height:38px;border-radius:8px;background:var(--accent-wash);display:flex;align-items:center;justify-content:center;color:var(--accent);flex:none}
    .report .icon svg{width:19px;height:19px}
    .chip{font-size:11px;font-weight:600;color:var(--ink-2);background:var(--page);border:1px solid var(--border);border-radius:5px;padding:2px 8px;letter-spacing:.04em}
    .report h3{font-size:15px;font-weight:600;letter-spacing:-.01em}
    .report p{font-size:13.5px;color:var(--ink-2);flex:1}
    .btn{display:inline-flex;align-items:center;justify-content:center;gap:8px;padding:9px 15px;border-radius:8px;
         background:var(--accent);color:#fff;font-size:13.5px;font-weight:600;text-decoration:none;border:none;cursor:pointer;
         transition:background .15s}
    .btn:hover{background:var(--accent-hover)}
    .btn svg{width:15px;height:15px}

    .master{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:22px 24px;box-shadow:var(--shadow);
            display:flex;align-items:center;justify-content:space-between;gap:20px;flex-wrap:wrap}
    .master .info{display:flex;align-items:flex-start;gap:15px}
    .master .icon{width:42px;height:42px;border-radius:9px;background:var(--accent-wash);display:flex;align-items:center;justify-content:center;color:var(--accent);flex:none}
    .master .icon svg{width:20px;height:20px}
    .master h3{font-size:15.5px;font-weight:650;letter-spacing:-.01em}
    .master p{font-size:13.5px;color:var(--ink-2);margin-top:2px;max-width:60ch}

    footer{border-top:1px solid var(--border);margin-top:52px;padding-top:20px;display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap;font-size:13px;color:var(--ink-3)}
    footer nav{display:flex;gap:20px}
    footer a{color:var(--ink-2);text-decoration:none;font-weight:500}
    footer a:hover{color:var(--accent)}
</style></head>
<body>
<header><div class="head-inner">
    <div class="brand">
        <div class="brand-mark">
            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round">
                <path d="M3 12h4l2.5-6.5 5 13L17 12h4"/>
            </svg>
        </div>
        <div>
            <div class="name">Infrastructure Reports</div>
            <div class="sub">Network Operations &middot; Zabbix</div>
        </div>
    </div>
    <div class="head-meta">
        <span class="live"><span class="dot"></span>Live data</span>
        <span id="today"></span>
    </div>
</div></header>

<div class="wrap">
    <div class="hero">
        <h1>Monitoring &amp; asset reports</h1>
        <p>Point-in-time Excel exports generated live from the Zabbix API &mdash; asset inventory, active problems, SLA compliance and availability.</p>
    </div>

    <div class="kpis">
        <div class="kpi">
            <div class="label">Hosts monitored</div>
            <div class="value" id="k-hosts">&ndash;</div>
            <div class="note">Enabled in Zabbix</div>
        </div>
        <div class="kpi">
            <div class="label">Active problems</div>
            <div class="value" id="k-problems">&ndash;</div>
            <div class="note">All severities, right now</div>
        </div>
        <div class="kpi">
            <div class="label"><span class="sdot" style="background:var(--critical)"></span>High &amp; disaster</div>
            <div class="value" id="k-severe">&ndash;</div>
            <div class="note">Problems needing attention</div>
        </div>
    </div>

    <h2 class="section">Standard reports</h2>
    <div class="reports">
        <div class="report">
            <div class="top">
                <div class="icon"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8"><rect x="3" y="4" width="18" height="7" rx="1.5"/><rect x="3" y="13" width="18" height="7" rx="1.5"/><path d="M7 7.5h.01M7 16.5h.01"/></svg></div>
                <span class="chip">XLSX</span>
            </div>
            <h3>Asset inventory</h3>
            <p>Hardware and software inventory for every monitored host: OS, serial numbers, MAC addresses, IPs and host groups.</p>
            <a class="btn" href="report/inventory" download><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 3v12m0 0l-4-4m4 4l4-4M4 19h16"/></svg>Download</a>
        </div>
        <div class="report">
            <div class="top">
                <div class="icon"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linejoin="round"><path d="M12 4l9 16H3l9-16z"/><path d="M12 10v4m0 3h.01" stroke-linecap="round"/></svg></div>
                <span class="chip">XLSX</span>
            </div>
            <h3>Active problems</h3>
            <p>Point-in-time export of every open problem, with host, description, severity and when it started.</p>
            <a class="btn" href="report/problems" download><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 3v12m0 0l-4-4m4 4l4-4M4 19h16"/></svg>Download</a>
        </div>
        <div class="report">
            <div class="top">
                <div class="icon"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linejoin="round"><path d="M12 3l8 3v6c0 4.5-3.2 7.6-8 9-4.8-1.4-8-4.5-8-9V6l8-3z"/><path d="M9 12l2 2 4-4" stroke-linecap="round"/></svg></div>
                <span class="chip">XLSX</span>
            </div>
            <h3>SLA compliance</h3>
            <p>Uptime against agreed service-level targets, per host, with pass/fail status for the reporting period.</p>
            <a class="btn" href="report/sla" download><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 3v12m0 0l-4-4m4 4l4-4M4 19h16"/></svg>Download</a>
        </div>
        <div class="report">
            <div class="top">
                <div class="icon"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M3 12h4l2.5-7 5 14L17 12h4"/></svg></div>
                <span class="chip">XLSX</span>
            </div>
            <h3>Availability detail</h3>
            <p>Host reachability deep-dive, including error detail for devices currently unreachable.</p>
            <a class="btn" href="report/availability" download><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 3v12m0 0l-4-4m4 4l4-4M4 19h16"/></svg>Download</a>
        </div>
        <div class="report">
            <div class="top">
                <div class="icon"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round"><path d="M5 20V10m7 10V4m7 16v-7"/></svg></div>
                <span class="chip">XLSX</span>
            </div>
            <h3>Top 100 triggers</h3>
            <p>The most frequently firing alerts over the last 7 days &mdash; the starting point for reducing alert noise.</p>
            <a class="btn" href="report/top-triggers" download><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 3v12m0 0l-4-4m4 4l4-4M4 19h16"/></svg>Download</a>
        </div>
    </div>

    <h2 class="section">Consolidated</h2>
    <div class="master">
        <div class="info">
            <div class="icon"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linejoin="round"><path d="M21 8l-9-5-9 5 9 5 9-5z"/><path d="M3 8v8l9 5 9-5V8"/><path d="M12 13v8"/></svg></div>
            <div>
                <h3>Full infrastructure report</h3>
                <p>All of the above in a single multi-sheet workbook: inventory, active problems, SLA compliance and availability detail.</p>
            </div>
        </div>
        <a class="btn" href="report/master" download><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 3v12m0 0l-4-4m4 4l4-4M4 19h16"/></svg>Download workbook</a>
    </div>

    <footer>
        <span>Data is queried live from the Zabbix API at download time.</span>
        <nav>
            <a href="/zabbix/">Zabbix</a>
            <a href="/grafana/">Grafana</a>
            <a href="/glpi/">GLPI</a>
        </nav>
    </footer>
</div>

<script>
document.getElementById('today').textContent = new Date().toLocaleDateString(undefined,
    {weekday:'short', year:'numeric', month:'short', day:'numeric'});

fetch('api/stats').then(function(r){ return r.json(); }).then(function(s){
    var fmt = function(v){ return (v === null || v === undefined) ? '\\u2013' : Number(v).toLocaleString(); };
    document.getElementById('k-hosts').textContent = fmt(s.hosts);
    document.getElementById('k-problems').textContent = fmt(s.problems);
    document.getElementById('k-severe').textContent = fmt(s.severe);
}).catch(function(){ /* tiles keep their dash placeholders */ });

// Report generation queries Zabbix live and can take a few seconds; give feedback.
document.querySelectorAll('a.btn').forEach(function(a){
    a.addEventListener('click', function(){
        var original = a.innerHTML;
        a.innerHTML = 'Preparing\\u2026';
        a.style.pointerEvents = 'none';
        setTimeout(function(){ a.innerHTML = original; a.style.pointerEvents = ''; }, 6000);
    });
});
</script>
</body></html>"""

@app.route('/')
def index(): return HTML

@app.route('/api/stats')
def api_stats():
    def as_int(v):
        try: return int(v)
        except (TypeError, ValueError): return None
    return jsonify({
        "hosts": as_int(zabbix.call("host.get", {"countOutput": True, "monitored_hosts": True})),
        "problems": as_int(zabbix.call("problem.get", {"countOutput": True})),
        "severe": as_int(zabbix.call("problem.get", {"countOutput": True, "severities": [4, 5]})),
    })

@app.route('/health')
def health(): return jsonify({"status": "healthy"}), 200

@app.route('/report/inventory')
def report_inventory():
    hosts = zabbix.get_detailed_inventory()
    data = []
    for h in hosts:
        inv = h.get('inventory', {})
        if isinstance(inv, list): inv = {} # Fix for empty inventory as list
        data.append({
            'Hostname': h.get('name', h.get('host')),
            'IP Address': (h.get('interfaces', [{}])[0].get('ip', 'N/A') if h.get('interfaces') else 'N/A'),
            'OS': inv.get('os', 'N/A'),
            'Hardware': inv.get('hardware', 'N/A'),
            'Serial Number': inv.get('serialno_a', 'N/A'),
            'MAC Address': inv.get('macaddress_a', 'N/A'),
            'Host Groups': ', '.join([g.get('name', '') for g in h.get('groups', [])]),
            'Monitoring Status': 'Enabled' if h.get('status') == '0' else 'Disabled'
        })
    return send_file(make_excel(data, 'Inventory'), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'asset_inventory_{datetime.now():%Y%m%d}.xlsx')

@app.route('/report/hosts')
def report_hosts(): return report_inventory()

@app.route('/report/problems')
def report_problems():
    probs = zabbix.get_active_problems()
    sev = {0: 'Not classified', 1: 'Info', 2: 'Warning', 3: 'Average', 4: 'High', 5: 'Disaster'}
    data = [{'Host': (p.get('hosts', [{}])[0].get('name', '?') if p.get('hosts') else '?'), 'Problem': p.get('description'), 'Severity': sev.get(int(p.get('priority', 0)), '?'), 'Time': p.get('lastchange')} for p in probs]
    return send_file(make_excel(data, 'Problems'), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'problems_report_{datetime.now():%Y%m%d}.xlsx')

@app.route('/report/sla')
def report_sla():
    sla_data = zabbix.get_native_slas()
    if not sla_data:
        avail_data = zabbix.get_availability_data()
        sla_data = [{'Host': item['Host'], 'SLA Target': '99.9%', 'Actual': item['Estimated Uptime'], 'Status': 'OK' if item['Estimated Uptime'] == '100%' else 'FAIL'} for item in avail_data]
    return send_file(make_excel(sla_data, 'SLA Report'), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'sla_report_{datetime.now():%Y%m%d}.xlsx')

@app.route('/report/availability')
def report_availability():
    data = zabbix.get_availability_data()
    return send_file(make_excel(data, 'Availability'), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'availability_report_{datetime.now():%Y%m%d}.xlsx')

@app.route('/report/top-triggers')
def report_top_triggers():
    data = zabbix.get_top_triggers()
    return send_file(make_excel(data, 'Top 100 Triggers'), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'top_100_triggers_{datetime.now():%Y%m%d}.xlsx')

@app.route('/report/master')
def report_master():
    wb = Workbook()
    # Sheet 1: Inventory
    inv_hosts = zabbix.get_detailed_inventory()
    inv_data = []
    for h in inv_hosts:
        inv = h.get('inventory', {})
        if isinstance(inv, list): inv = {}
        inv_data.append({'Hostname':h.get('name'),'IP':(h.get('interfaces',[{}] )[0].get('ip') if h.get('interfaces') else 'N/A'),'OS':inv.get('os','N/A'),'Status':'Enabled' if h.get('status')=='0' else 'Disabled'})
    ws1 = wb.active; ws1.title = "Detailed Inventory"; format_excel_sheet(ws1, inv_data, "Inventory")
    # Sheet 2: Problems
    probs = zabbix.get_active_problems(); sev = {0:'Not classified',1:'Info',2:'Warning',3:'Average',4:'High',5:'Disaster'}
    p_data = [{'Host':(p.get('hosts',[{}])[0].get('name','?') if p.get('hosts') else '?'),'Problem':p.get('description'),'Severity':sev.get(int(p.get('priority',0)),'?'),'Time':p.get('lastchange')} for p in probs]
    ws2 = wb.create_sheet("Active Problems"); format_excel_sheet(ws2, p_data, "Problems")
    # Sheet 3: SLA
    avail_data = zabbix.get_availability_data()
    sla_data = [{'Host': item['Host'], 'SLA Target': '99.9%', 'Actual': item['Estimated Uptime'], 'Status': 'OK' if item['Estimated Uptime'] == '100%' else 'FAIL'} for item in avail_data]
    ws3 = wb.create_sheet("SLA Compliance"); format_excel_sheet(ws3, sla_data, "SLA")
    # Sheet 4: Availability
    ws4 = wb.create_sheet("Availability Detail"); format_excel_sheet(ws4, avail_data, "Availability")

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'Zabbix_Full_Infrastructure_Report_{datetime.now():%Y%m%d}.xlsx')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
